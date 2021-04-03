#Vamos a importar la librería de excel
import openpyxl

#Crea un archivo de excel
deimer= openpyxl.Workbook()

hojad = deimer.active

# Voy a escribir en la hoja de cálculo
hojad['A1'] = "Canciones"
hojad['A2'] = "Vamos a cantar"
hojad['B1'] = "Género"
hojad['B2'] = "¿Qué canción?"
hojad['C1'] = "Año"
hojad['C2'] ="Pues yo no sé, dime tú"
hojad['D1'] = "¿Es famosa la canción?"
hojad['D2'] = "Ok!... entonces cantaremos los pollitos"

deimer.save('excel_deimer.xlsx')

#Ahora vamos a leer el archivos
deimer = openpyxl.load_workbook('excel_deimer.xlsx')

#value nos permite visualizar lo que se encuentra en dicha fila 
print(hojad['A1'].value)
print(hojad['B1'].value)
print(hojad['C1'].value)
print(hojad['D1'].value)

#volvemos a abrir el archivo
deimer = openpyxl.load_workbook('excel_deimer.xlsx')

#Crea otra hoja de calculo en el mismo archivo de excel
canciones=deimer.create_sheet('canciones')
bailes=deimer.create_sheet('bailes',0)
discotecas=deimer.create_sheet('discotecas',-1)

deimer.save('excel_deimer.xlsx')